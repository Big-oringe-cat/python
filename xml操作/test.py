#! /usr/bin/python
import xml.etree.ElementTree as ET
import os
import sys
reload(sys)
sys.setdefaultencoding('utf8')
from openpyxl import load_workbook

wb = load_workbook('scf.xlsx')
sheet = wb.active
a1=sheet['A1']

dict = {}
for i in sheet.rows:
    name = str(i[1].value)
    list = []
    for count in range(sheet.max_column):
        list.append(str(i[count].value))
    dict[name] = list


tree = ET.parse("zbx_export_templates.xml")
root = tree.getroot()
for i in root[-1].findall("trigger"):
    name = ""
    for j in i.findall("name"):
        name = str(j.text)
    for j in i.findall("url"):
        if dict.has_key(name):
            j.text = dict[name][2]
        else:
            print name
    for j in i.findall("description"):
        if dict.has_key(name):
            j.text = dict[name][0] + '\n' + dict[name][-1]
    
tree2 = ET.ElementTree(root)
tree2.write("note.xml", encoding="utf-8", xml_declaration=True)
